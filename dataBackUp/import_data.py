import os
import django
import pandas as pd
from django.db import IntegrityError

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'database_jkm_wyx_yzh.settings')
django.setup()

'''
Major_type = (
    (1, '哲学'),
    (2, '经济学'),
    (3, '法学'),
    (4, '教育学'),
    (5, '文学'),
    (6, '历史学'),
    (7, '理学'),
    (8, '工学'),
    (9, '农学'),
    (10, '医学'),
    (11, '军事学'),
    (12, '管理学'),
    (13, '艺术学'),
    (14, '交叉学科'),
)
'''


def getType(door):
    if door == "哲学":
        return 1
    if door == "经济学":
        return 2
    if door == "法学":
        return 3
    if door == "教育学":
        return 4
    if door == "文学":
        return 5
    if door == "历史学":
        return 6
    if door == "理学":
        return 7
    if door == "工学":
        return 8
    if door == "农学":
        return 9
    if door == "医学":
        return 10
    if door == "军事学":
        return 11
    if door == "管理学":
        return 12
    if door == "艺术学":
        return 13
    if door == "交叉学科":
        return 14
    return 0


def insertMajorCatalog():
    from app.models import Major, ExamInfo, University, MajorUniversity
    Major.objects.all().delete()
    ExamInfo.objects.all().delete()
    University.objects.all().delete()
    MajorUniversity.objects.all().delete()
    form = pd.read_excel(r'.\专业目录\24考研招生目录.xlsx')  #
    count = 0
    for index, row in form.iterrows():
        data = form.iloc[index].values
        title = form.columns.values
        tmp = zip(title, data)
        dic = dict(tmp)
        major_id = dic['专业代码']  # 专业代码
        major_class = getType(dic['门类'])  # 门类
        major = dic['专业']
        major_name = major[major.find(")") + 1:]
        university = dic['招生单位']
        university_id = dic['招生单位'].split(')')[0][1::]
        university_name = university[university.find(")") + 1:]
        location = dic['所在地']
        school = dic['院系所'][dic['院系所'].find(")") + 1:]
        # 业务课一、业务课二、外语、政治
        professional_exam1 = dic['业务课一']
        professional_exam2 = dic['业务课二']
        english_exam = dic['外语']
        politic_exam = dic['政治']
        try:
            if not University.objects.filter(university_id=university_id).exists():
                # print(university_name)
                inputU = University(university_id=university_id, name=university_name, location=location)
                inputU.save()
            if not Major.objects.filter(major_id=major_id).exists():
                # print(major_name)
                inputM = Major(major_id=major_id, major_name=major_name, major_type=major_class)
                inputM.save()
            if not MajorUniversity.objects.filter(major_id=major_id, university_id_id=university_id).exists():
                inputMU = MajorUniversity(major_id=major_id, university_id_id=university_id, school=school)
                inputMU.save()
                inputE = ExamInfo(major_id_id=major_id, university_id_id=university_id,
                                  professional_exam1=professional_exam1,
                                  professional_exam2=professional_exam2, english_exam=english_exam,
                                  politic_exam=politic_exam)
                inputE.save()
                print("insert " + str(count))
                count = count + 1
        except IntegrityError as e:
            print(e)
    print("major finished")


def getWebs():
    import requests
    from lxml import etree
    import xlsxwriter
    workbook = xlsxwriter.Workbook('gw.xlsx')
    worksheet = workbook.add_worksheet()
    gwCount = 0
    base_url = 'http://u.feelingmsg.com/u/'
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36"
    }
    response = requests.get(headers=headers, url=base_url + 'menu.php')
    parser = etree.HTMLParser(encoding='utf-8')
    tree = etree.HTML(response.content, parser=parser)
    r = tree.xpath('/html/body/table[4]')[0]
    for tr in r.xpath('.//tr'):
        try:  # 天津没了
            district_url = base_url + tr.xpath('./td//a/@href')[0]
            response0 = requests.get(headers=headers, url=district_url)
            response0.encoding = 'GBK'
            tree0 = etree.HTML(response0.text)
            u_block = tree0.xpath('/html/body/table[4]')[0]
            i = 0
            for u_tr in u_block.xpath('.//tr'):  # 一行
                if (i == 0):
                    i = 1
                    continue
                try:
                    for u_td in u_tr.xpath('./td'):
                        gw_url = u_td.xpath('.//a/@href')[0]
                        gw_name = u_td.xpath('.//a/text()')[0]
                        print(gw_name + ":" + gw_url)
                        worksheet.write('A' + str(gwCount), gw_name)
                        worksheet.write('B' + str(gwCount), gw_url)
                        gwCount = gwCount + 1
                except IndexError as e:
                    break
        except IndexError:
            print('city not found')
    workbook.close()


def insertUniversityInfo():
    from app.models import University
    import requests
    import openpyxl

    getWebs() #爬官网
    workbookG = openpyxl.load_workbook('./大学信息/gw.xlsx')
    sheetG = workbookG.active

    workbookL = openpyxl.load_workbook('./大学信息/label.xlsx')
    sheetL = workbookL.active
    base_url = 'https://t1.chei.com.cn/common/xh/'
    for row in University.objects.all():
        # 校徽
        school_id = str(row.university_id)
        url = base_url + school_id + ".jpg"
        response = requests.get(url)
        with open('/static/img/university' + school_id + '.jpg', 'wb') as f:
            f.write(response.content)
        row.image = '/static/img/university/' + school_id + '.jpg'
        row.save()

        # 联系方式
        school_name = row.name
        for cell in sheetG['A']:
            if cell.value == school_name:
                gw = sheetG.cell(row=cell.row, column=2).value
                row.info = gw
                break
        label = 0
        for cell in sheetL['B']:
            if cell.value == school_name:
                if sheetL.cell(row=cell.row, column=4).value == '是':
                    label = 2
                    # print(label)
                elif sheetL.cell(row=cell.row, column=5).value == '是':
                    label = 1
                    # print(label)
                break
        row.university_tag = label
        row.save()

    workbookG.close()
    workbookL.close()


#
# total_score = models.DecimalField(verbose_name='总分', max_digits=6, decimal_places=2)
#   score1 = models.IntegerField(verbose_name='初试分数')
#   score2 = models.DecimalField(verbose_name='复试分数', max_digits=6, decimal_places=2)
#   year = models.IntegerField(verbose_name='年份')
#   major_id = models.ForeignKey(to='Major', to_field='major_id', on_delete=models.CASCADE)
#   university_id = models.ForeignKey(to='University', to_field='university_id', on_delete=models.CASCADE)


def insertAdmissions():
    from app.models import AdmitInfo,Major

    # AdmitInfo.objects.all().delete()
    #
    university_id = 10006
    year = 2019
    major_id = '081000'
    form = pd.read_excel(r'.\录取信息\化学学院2022年硕士研究生一志愿复试结果公示.xlsx')  #
    for index, row in form.iterrows():
        data = form.iloc[index].values
        title = form.columns.values
        tmp = zip(title, data)
        dic = dict(tmp)
        total_score = dic['总成绩']
        score1 = dic['初试成绩']
        score2 = dic['复试成绩']
        print(total_score,score1,score2)
        inputA = AdmitInfo(total_score=total_score,year=year,major_id_id=major_id,university_id_id=university_id,score1=score1,score2=score2)
        inputA.save()
    print('info finish')


if __name__ == '__main__':
    # 专业库
    insertMajorCatalog()
    # 大学补充信息 电话 邮箱 头像
    insertUniversityInfo()
    # 录取信息
    insertAdmissions()
